import docx
import pptx
import pathlib
import openpyxl
import urllib.parse
import pebble.concurrent

import windows
import utilities


@pebble.concurrent.process(timeout=300)
def move_xls(path_source: str, path_destination: str) -> None:
    with windows.open_xls(path_source) as file:
        file.SaveAs(path_destination, FileFormat=51)
    openpyxl.load_workbook(path_destination)


@pebble.concurrent.process(timeout=300)
def move_doc(path_source: str, path_destination: str) -> None:
    with windows.open_doc(path_source) as file:
        file.SaveAs(path_destination, FileFormat=16)
    docx.Document(path_destination)


@pebble.concurrent.process(timeout=300)
def move_ppt(path_source: str, path_destination: str) -> None:
    with windows.open_ppt(path_source) as file:
        file.SaveAs(path_destination, FileFormat=24)
    pptx.Presentation(path_destination)


def move(full: bool, directory_src: pathlib.Path, directory_dst: pathlib.Path, watch: utilities.Watch) -> None:

    suffix_to_function_suffix = {".xlsx": (move_xls, ".xlsx"), ".xls": (move_xls, ".xlsx"), ".docx": (move_doc, ".docx"), ".doc": (move_doc, ".docx"), ".pptx": (move_ppt, ".pptx"), ".ppt": (move_ppt, ".pptx")}

    path_idx = [pathlib.Path(*path.parts[1:]) for path in directory_src.glob("*") if path.suffix == ".xlsx" and not path.name.startswith(".")][0]

    paths_dir = {pathlib.Path(*path.parts[1:]): [] for path in directory_src.rglob("*") if path.suffix in (".xlsx", ".xls", ".docx", ".doc", ".pptx", ".ppt") and not path.name.startswith(".")}
    paths_idx = {}

    workbook = openpyxl.load_workbook(str(directory_src / path_idx))
    for sheetname, (rows, cols) in utilities.worksheets_dimensions(str(directory_src / path_idx)).items():
        for row in range(1, rows + 2):
            for col in range(1, cols + 1):
                cell = workbook[sheetname].cell(row, col)
                if cell.hyperlink is not None:
                    hyperlink_old = pathlib.Path(urllib.parse.unquote(cell.hyperlink.target))
                    hyperlink_src = directory_src / hyperlink_old
                    if hyperlink_src.exists() and ".." not in hyperlink_src.parts:
                        if hyperlink_src.is_file():
                            hyperlink = hyperlink_src
                            path = pathlib.Path(*hyperlink.parts[1:])
                            if path.suffix in (".xlsx", ".xls", ".docx", ".doc", ".pptx", ".ppt") and not path.name.startswith("."):
                                if path not in paths_idx:
                                    paths_idx[path] = [cell]
                                else:
                                    paths_idx[path].append(cell)
                        if hyperlink_src.is_dir():
                            for hyperlink in hyperlink_src.glob("*"):
                                path = pathlib.Path(*hyperlink.parts[1:])
                                if path.suffix in (".xlsx", ".xls", ".docx", ".doc", ".pptx", ".ppt") and not path.name.startswith("."):
                                    if path not in paths_idx:
                                        paths_idx[path] = []

    paths = paths_idx if not full else paths_dir | paths_idx

    watch.beginning(len(paths))
    for index, (path, cells) in enumerate(paths.items(), start=1):
        watch.current(index, path)
        watch.print("moving")
        try:
            function, suffix = suffix_to_function_suffix[path.suffix]
            path_old = path
            path_new = path.with_suffix(suffix)
            path_src = directory_src / path_old
            path_dst = directory_dst / path_new
            path_dst.parent.mkdir(parents=True, exist_ok=True)
            function(str(path_src.resolve()), str(path_dst.resolve())).result()
        except Exception as exception:
            path_dst.unlink(missing_ok=True)
            for cell in cells: cell.hyperlink.target = str(path_old)
            watch.print(f"{exception}")
        else:
            for cell in cells: cell.hyperlink.target = str(path_new)
            watch.print("moved")

    workbook.save(str(directory_dst / path_idx))
